import time  
import mylabnew as mll  
import pandas as pd  
import openpyxl  
from pulp import LpProblem, LpMinimize, LpVariable, lpSum, value

start_time = time.time()  

path1 = 'Raw 1600 Dmus DataSet.xlsx'  
path2 = 'Raw 1400 Dmus DataSet ( no Zero ).xlsx'  
path3 = 'test.xlsx'  

file_name = path2  

drive = "d:\\"  

df = pd.read_excel(f"{drive}{file_name}", sheet_name='data')  

# ------------  finding the number of inputs and outputs --------------------  
t = 0  
for i in df.columns:  
    t += 1  
    if i == "O1":  
        I = t - 2  
        O = len(df.columns) - I - 1  

# ------------- getting the Dimentions of dataFrame --------------------------  
print("\n your matrix id : {} * {} \n".format(len(df.index), len(df.columns)))  
print("\n Your dataset is : \n", df)  

# ------------- finding the Iedal DMU ----------------------------------------  
Idmu = mll.find_IDMU(df)  
print(" \n Ideal DMU is : \n {} ".format(Idmu))  

# ------------- finding DMUI and DMUO  ----------------------------------------  
# -----------------------DMUOI-------------------------------------------------  
setdmui = mll.filter_dmuo_dmui_sets(df, I, O)[0]  
print("\nDMU I selection:\n", setdmui)  
setdmuo = mll.filter_dmuo_dmui_sets(df, I, O)[1]  
print("\nDMU O selection:\n", setdmuo)  

# ایجاد یک فایل اکسل جدید  
try:  
    with pd.ExcelWriter(f"{drive}_UNIQ INDEXES _{file_name}", engine='openpyxl') as writer:  
        setdmui.to_excel(writer, sheet_name='input', index=False)  
        setdmuo.to_excel(writer, sheet_name='output', index=False)  
except PermissionError:  
    print("دسترسی به فایل اکسل مجاز نیست. لطفا فایل را ببندید و دوباره امتحان کنید.")  

s_Dmu = pd.concat([setdmui, setdmuo], ignore_index=True)  
print("Selected Dmus for Process :\n", s_Dmu)  

df = df.drop(s_Dmu.index)  
print(df)  


from openpyxl.styles import Alignment, Font  
from openpyxl.utils.dataframe import dataframe_to_rows  
from openpyxl.workbook import Workbook  
from openpyxl import load_workbook  

K = df['DMU'].tolist()  
I = [col for col in df.columns if col.startswith('I')]  
J = [col for col in df.columns if col.startswith('O')]  
eps = 0.000001  

# Parameters building  
X = {  
    i: {  
        k: 0 for k in K  
    } for i in I  
}  
Y = {  
    j: {  
        k: 0 for k in K  
    } for j in J  
}


KK = s_Dmu['DMU'].tolist()  
II = [col for col in s_Dmu.columns if col.startswith('I')]  
JJ = [col for col in s_Dmu.columns if col.startswith('O')]  

# Parameters building  
XX = {  
    i1: {  
        k1: 0 for k1 in KK  
    } for i1 in II  
}  
YY = {  
    j1: {  
        k1: 0 for k1 in KK  
    } for j1 in JJ  
}  


for i in II:  
    for k in KK:  
        XX[i][k] = s_Dmu[i][k]  
for j in JJ:  
    for k in KK:  
        YY[j][k] = s_Dmu[j][k]  

for i in I:  
    for k in K:  
        X[i][k] = df[i][k]  
for j in J:  
    for k in K:  
        Y[j][k] = df[j][k]  

results = []  
for dmu in KK:  
    start_time = time.time()  

    # Model building  
    model = LpProblem(f'DEA_{dmu}', LpMinimize)  
    lambda_k = {k: LpVariable(f'lambda_{k}', lowBound=0) for k in KK}  

    # Objective Function setting  
    allfa_r = LpVariable('allfa_r', lowBound=0)  
    model += allfa_r  

    # Constraints setting  
    for i in II:  
        model += lpSum([  
            allfa_r - (lambda_k[k] * XX[i][k])  
        ] for k in KK) >= - XX[i][dmu]  
    for j in JJ:  
        model += lpSum([  
            allfa_r + lambda_k[k] * YY[j][k]  
        ] for k in KK) >= YY[j][dmu]  
    model += lpSum([lambda_k[k] for k in KK]) == 1  # Added convexity constraint  

    # Model solving  
    model.solve()  

    # Collect results  
    efficiency = value(model.objective)  
    result = {  
        'DMU': dmu,  
        'Efficiency': round(efficiency, 3),  
        'Time (s)': round(time.time() - start_time, 3),  
       
    }  
    for k in KK:  
        result[f'l_{k}'] = round(value(lambda_k[k]), 3)  

    results.append(result)  

df_results = pd.DataFrame(results)  

sheet_name_results = 'Dominated Dmus'  
sheet_name_data = 'data'  

try:  
    wb = load_workbook(f"{drive}_ DOMINATED BY UNIQ DMUS _{file_name}")  
    if 'Dominated Dmus' in wb.sheetnames:  
        sheet_name_results = f'Dominated Dmus_{len(wb["Dominated Dmus"])+1}'  
    ws_results = wb.create_sheet(sheet_name_results)  
    ws_data = wb['data']  

    # Write results to the new sheet  
    for row in dataframe_to_rows(df_results, index=False, header=True):  
        ws_results.append(row)  

    # Adjust column widths and center-align the content  
    for col in range(1, len(df_results.columns) + 1):  
        cell = ws_results.cell(row=1, column=col)  
        cell.font = Font(bold=True)  
        cell.alignment = Alignment(horizontal='center')  
        ws_results.column_dimensions[chr(col + 64)].width = max(len(str(cell.value)), *[len(str(x)) for x in df_results[df_results.columns[col - 1]]]) + 2  
    for row in range(2, ws_results.max_row + 1):  
        for col in range(1, len(df_results.columns) + 1):  
            cell = ws_results.cell(row=row, column=col)  
            cell.alignment = Alignment(horizontal='center')  

    # Save the workbook  
    wb.save(f"{drive}_ DOMINATED BY UNIQ DMUS _{file_name}")  

except Exception as e:  
    print(f"Error occurred: {e}")  
    # Create a new Excel file and write the header  
    wb = Workbook()  
    ws_results = wb.active  
    ws_results.title = 'Dominated Dmus'  
    ws_data = wb.create_sheet('data')  

    # Write data to the sheets  
    for row in dataframe_to_rows(df, index=False, header=True):  
        ws_data.append(row)  
    for row in dataframe_to_rows(df_results, index=False, header=True):  
        ws_results.append(row)  

    # Adjust column widths and center-align the content  
    for ws in [ws_results, ws_data]:  
        for col in range(1, len(df.columns) + 1):  
            cell = ws.cell(row=1, column=col)  
            cell.font = Font(bold=True)  
            cell.alignment = Alignment(horizontal='center')  
            ws.column_dimensions[chr(col + 64)].width = max(len(str(cell.value)), *[len(str(x)) for x in df[df.columns[col - 1]]]) + 2  
        for row in range(2, ws.max_row + 1):  
            for col in range(1, len(df.columns) + 1):  
                cell = ws.cell(row=row, column=col)  
                cell.alignment = Alignment(horizontal='center')  

    wb.save(f"{drive}_ DOMINATED BY UNIQ DMUS _{file_name}")