from pulp import LpProblem, LpMinimize, LpVariable, lpSum, value  
import pandas as pd  
import time  
from openpyxl.styles import Alignment, Font  
from openpyxl.utils.dataframe import dataframe_to_rows  
from openpyxl.workbook import Workbook  
from openpyxl import load_workbook  # Import load_workbook  

path = 'd:\\r1.xlsx'  

# Set building  
df = pd.read_excel(path, sheet_name='data')  
K = df['DMU'].tolist()  
I = [col for col in df.columns if col.startswith('I')]  
J = [col for col in df.columns if col.startswith('O')]  
eps = 0.000001  

# Parameters building  
X = {  
    i: {  
        k: 0 for k in K  
    } for i in I  
}  
Y = {  
    j: {  
        k: 0 for k in K  
    } for j in J  
}  

# Populate input and output parameters  
for i in I:  
    for k in K:  
        X[i][k] = df.loc[df['DMU'] == k, i].values[0]  
for j in J:  
    for k in K:  
        Y[j][k] = df.loc[df['DMU'] == k, j].values[0]  

# ε-constrained DEA Model  
results = []  
for dmu in K:  
    # Start timer  
    start_time = time.time()  

    # Model Building  
    model = LpProblem('ε-constrained_model', LpMinimize)  

    # Decision variables Building  
    theta_r = LpVariable(f'theta_r')  
    lambda_k = {k: LpVariable(f'lambda_{k}', lowBound=0) for k in K}  
    s_i = {i: LpVariable(f's_i_{i}', lowBound=0) for i in I}  
    s_j = {j: LpVariable(f's_j_{j}', lowBound=0) for j in J}  

    # Objective Function setting  
    model += theta_r + eps * (lpSum(s_i.values()) + lpSum(s_j.values()))  

    # Constraints setting  
    for i in I:  
        model += lpSum([  
            lambda_k[k] * X[i][k]  
        ] for k in K) + s_i[i] == theta_r * X[i][dmu]  
    for j in J:  
        model += lpSum([  
            lambda_k[k] * Y[j][k]  
        ] for k in K) - s_j[j] >= Y[j][dmu]  
    model += lpSum([lambda_k[k] for k in K]) == 1  # Added convexity constraint  

    # Model solving  
    model.solve()  

    # Collect results  
    efficiency = value(model.objective) - eps * (value(lpSum(s_i.values())) + value(lpSum(s_j.values())))  
    result = {  
        'DMU': dmu,  
        'Efficiency': round(efficiency, 3),  
        'Time (s)': round(time.time() - start_time, 3)  
    }  
    for k in K:  
        result[f'lambda_{k}'] = round(value(lambda_k[k]), 3)  
    for i in I:  
        result[f'slack_i_{i}'] = round(value(s_i[i]), 3)  
    for j in J:  
        result[f'slack_i_{j}'] = round(value(s_j[j]), 3)  
    result['epsilon'] = round(eps * (value(lpSum(s_i.values())) + value(lpSum(s_j.values()))), 9)  
    results.append(result)  

df_results = pd.DataFrame(results)  

filename = path  
sheet_name_results = 'result_ε_constrained_model'  
sheet_name_data = 'data'  

try:  
    wb = load_workbook(filename)  
    if 'result_ε_constrained_model' in wb.sheetnames:  
        sheet_name_results = f'result_ε_constrained_model_{len(wb["result_ε_constrained_model"])+1}'  
    ws_results = wb.create_sheet(sheet_name_results)  
    ws_data = wb['data']  
    
    # Write results to the new sheet  
    for row in dataframe_to_rows(df_results, index=False, header=True):  
        ws_results.append(row)  
        
    # Adjust column widths and center-align the content  
    for col in range(1, len(df_results.columns) + 1):  
        cell = ws_results.cell(row=1, column=col)  
        cell.font = Font(bold=True)  
        cell.alignment = Alignment(horizontal='center')  
        ws_results.column_dimensions[chr(col + 64)].width = max(len(str(cell.value)), *[len(str(x)) for x in df_results[df_results.columns[col - 1]]]) + 2  
    for row in range(2, ws_results.max_row + 1):  
        for col in range(1, len(df_results.columns) + 1):  
            cell = ws_results.cell(row=row, column=col)  
            cell.alignment = Alignment(horizontal='center')  
            
    # Save the workbook  
    wb.save(filename)  
    
except Exception as e:  
    print(f"Error occurred: {e}")  
    # Create a new Excel file and write the header  
    wb = Workbook()  
    ws_results = wb.active  
    ws_results.title = 'result_ε_constrained_model'  
    ws_data = wb.create_sheet('data')  
    
    # Write data to the sheets  
    for row in dataframe_to_rows(df, index=False, header=True):  
        ws_data.append(row)  
    for row in dataframe_to_rows(df_results, index=False, header=True):  
        ws_results.append(row)  
        
    # Adjust column widths and center-align the content  
    for ws in [ws_results, ws_data]:  
        for col in range(1, len(df.columns) + 1):  
            cell = ws.cell(row=1, column=col)  
            cell.font = Font(bold=True)  
            cell.alignment = Alignment(horizontal='center')  
            ws.column_dimensions[chr(col + 64)].width = max(len(str(cell.value)), *[len(str(x)) for x in df[df.columns[col - 1]]]) + 2  
        for row in range(2, ws.max_row + 1):  
            for col in range(1, len(df.columns) + 1):  
                cell = ws.cell(row=row, column=col)  
                cell.alignment = Alignment(horizontal='center')  
                
    wb.save(filename)