from pulp import LpProblem, LpMinimize, LpVariable, lpSum, value
import pandas as pd
import time
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

path = 'd:\\11.xlsx'

# Read data
df = pd.read_excel(path, sheet_name='data')
K = df['DMU'].tolist()
I = [col for col in df.columns if col.startswith('I')]
J = [col for col in df.columns if col.startswith('O')]
eps = 0.000001

# Prepare input and output parameters
X = {i: {k: df.loc[df['DMU'] == k, i].values[0] for k in K} for i in I}
Y = {j: {k: df.loc[df['DMU'] == k, j].values[0] for k in K} for j in J}

# Prepare to store cross-efficiency results
cross_efficiency = []

# Calculate DEA efficiency and weights for each DMU
dea_results = []
lambda_weights = {}

for dmu in K:
    # Create and solve DEA model
    model = LpProblem(f"DEA_Model_{dmu}", LpMinimize)
    
    # Decision variables
    theta = LpVariable(f"theta_{dmu}")
    lambda_k = LpVariable.dicts("lambda", K, lowBound=0)
    slack_i = LpVariable.dicts("slack_i", I, lowBound=0)
    slack_j = LpVariable.dicts("slack_j", J, lowBound=0)
    
    # Objective function
    model += theta + eps * (lpSum(slack_i[i] for i in I) + lpSum(slack_j[j] for j in J))
    
    # Input constraints
    for i in I:
        model += lpSum(lambda_k[k] * X[i][k] for k in K) + slack_i[i] == theta * X[i][dmu]
    
    # Output constraints
    for j in J:
        model += lpSum(lambda_k[k] * Y[j][k] for k in K) - slack_j[j] == Y[j][dmu]
    
    # Convexity constraint
    model += lpSum(lambda_k[k] for k in K) == 1
    
    # Solve the model
    start_time = time.time()
    model.solve()
    elapsed_time = time.time() - start_time
    
    # Store efficiency and weights
    efficiency = value(theta)
    dea_results.append({
        'DMU': dmu,
        'Efficiency': round(efficiency, 3),
        'Time (s)': round(elapsed_time, 3)
    })
    
    # Store lambda weights for cross-efficiency
    lambda_weights[dmu] = {k: value(lambda_k[k]) for k in K}

# Calculate cross-efficiency using other DMUs' weights
cross_efficiency = []
for dmu in K:
    avg_efficiency = 0
    count = 0
    
    for other_dmu in K:
        if other_dmu != dmu:
            # Use weights from other_dmu's DEA solution
            lambda_other = lambda_weights[other_dmu]
            
            # Calculate weighted output and input
            total_output = sum(Y[j][dmu] * lambda_other[k] for k in K for j in J)
            total_input = sum(X[i][dmu] * lambda_other[k] for k in K for i in I)
            
            if total_input != 0:
                efficiency = total_output / total_input
                avg_efficiency += efficiency
                count += 1
    
    if count > 0:
        cross_eff = round(avg_efficiency / count, 3)
    else:
        cross_eff = 0  # If all DMUs are identical
        
    cross_efficiency.append({
        'DMU': dmu,
        'Cross_Efficiency': cross_eff
    })

# Export cross-efficiency results to Excel
try:
    wb = load_workbook(filename=path)
    if 'Cross_Efficiency' in wb.sheetnames:
        ws = wb.create_sheet('Cross_Efficiency')
    else:
        ws = wb.create_sheet('Cross_Efficiency')
    
    # Write cross_efficiency data
    for row in dataframe_to_rows(pd.DataFrame(cross_efficiency), index=False, header=True):
        ws.append(row)
    
    # Adjust column widths and formatting
    for col in range(1, len(pd.DataFrame(cross_efficiency).columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col)].width = 15  # Adjust column width
        
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(pd.DataFrame(cross_efficiency).columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
    
    wb.save(path)
    
except Exception as e:
    print(f"Error occurred: {e}")
    # Create a new Excel file and write the results
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cross_Efficiency'
    
    for row in dataframe_to_rows(pd.DataFrame(cross_efficiency), index=False, header=True):
        ws.append(row)
    
    for col in range(1, len(pd.DataFrame(cross_efficiency).columns) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col)].width = 15  # Adjust column width
        
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(pd.DataFrame(cross_efficiency).columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            
    wb.save(path)